"""
core/exporters.py
-----------------
Export pipeline results to:
  - CSV (for Power BI)
  - Excel workbook (multi-sheet data)
  - Excel evidence report (formatted, human-readable, one row per excerpt)
  - JSON (D3 heatmap + network graph format)
  - Annotated evidence report (Markdown)
  - PDF evidence report (via reportlab)
"""

from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any, Dict, List

import pandas as pd

from core.search_engine import SearchResult


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def results_to_dataframe(results: List[SearchResult]) -> pd.DataFrame:
    """Convert SearchResult list to a flat DataFrame."""
    rows = []
    for r in results:
        rows.append(
            {
                "Document": r.doc_filename,
                "Page": r.page,
                "Theme": r.theme,
                "Excerpt": r.excerpt,
                "Keyword Hit": r.keyword_hit,
                "Matched Keyword": r.matched_keyword,
                "SBERT Score": round(r.sbert_score, 4),
                "CrossEncoder Score": round(r.cross_encoder_score, 4),
                "OpenAI Score": round(r.openai_score, 4),
                "Final Score": round(r.final_score, 4),
                "EO Relevance Score": round(getattr(r, "eo_relevance_score", 0.0), 4),
                "Validation Category": r.validation_category,
                "Human Label": r.human_label,
            }
        )
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

def to_csv_bytes(results: List[SearchResult]) -> bytes:
    """Returns UTF-8 encoded CSV bytes suitable for Power BI import."""
    df = results_to_dataframe(results)
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# Excel data export (multi-sheet)
# ---------------------------------------------------------------------------

def to_excel_bytes(results: List[SearchResult]) -> bytes:
    """
    Returns an Excel workbook with multiple sheets:
      - All Evidence, Valid Evidence, Weak Evidence, Not Relevant
      - Theme Summary pivot
      - Human Reviewed
    """
    df = results_to_dataframe(results)
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All Evidence", index=False)

        for cat, sheet_name in [
            ("VALID EVIDENCE", "Valid Evidence"),
            ("WEAK EVIDENCE", "Weak Evidence"),
            ("NOT RELEVANT", "Not Relevant"),
        ]:
            sub = df[df["Validation Category"] == cat]
            if not sub.empty:
                sub.to_excel(writer, sheet_name=sheet_name, index=False)

        if not df.empty:
            pivot = (
                df[df["Validation Category"].isin(["VALID EVIDENCE", "WEAK EVIDENCE"])]
                .groupby(["Document", "Theme"])
                .size()
                .unstack(fill_value=0)
                .reset_index()
            )
            pivot.to_excel(writer, sheet_name="Theme Summary", index=False)

        labelled = df[df["Human Label"] != ""]
        if not labelled.empty:
            labelled.to_excel(writer, sheet_name="Human Reviewed", index=False)

    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Excel evidence REPORT (formatted, human-readable)
# ---------------------------------------------------------------------------

def to_excel_report_bytes(results: List[SearchResult]) -> bytes:
    """
    Returns a formatted Excel report structured as a readable evidence document:
      - Cover sheet with summary statistics
      - One sheet per document (valid + weak evidence only)
      - Each excerpt shown as a formatted block with theme, score, page, text
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    df = results_to_dataframe(results)
    relevant = df[df["Validation Category"].isin(["VALID EVIDENCE", "WEAK EVIDENCE"])]

    wb = Workbook()

    # ---- Colour palette ---------------------------------------------------
    CLR_HEADER_BG   = "0A3D5C"   # dark blue header
    CLR_HEADER_FG   = "FFFFFF"
    CLR_VALID_BG    = "D4EDDA"   # green tint
    CLR_VALID_BADGE = "155724"
    CLR_WEAK_BG     = "FFF3CD"   # amber tint
    CLR_WEAK_BADGE  = "856404"
    CLR_ROW_ALT     = "F8F9FA"
    CLR_SECTION_BG  = "E8F4FD"   # light blue section header

    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def header_font(size=11, bold=True):
        return Font(name="Calibri", size=size, bold=bold, color=CLR_HEADER_FG)

    def body_font(size=10, bold=False, color="000000"):
        return Font(name="Calibri", size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill(fill_type="solid", fgColor=hex_color)

    def wrap_align(horizontal="left", vertical="top"):
        return Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

    # ---- Cover sheet -------------------------------------------------------
    ws_cover = wb.active
    ws_cover.title = "Summary"
    ws_cover.column_dimensions["A"].width = 35
    ws_cover.column_dimensions["B"].width = 20

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    valid_count = len(df[df["Validation Category"] == "VALID EVIDENCE"])
    weak_count  = len(df[df["Validation Category"] == "WEAK EVIDENCE"])
    docs        = df["Document"].nunique()
    themes      = df["Theme"].nunique()

    cover_rows = [
        ("EO Skills Policy Evidence Report", "", CLR_HEADER_BG, 16, True),
        (f"Generated: {ts}", "", None, 10, False),
        ("", "", None, 10, False),
        ("SUMMARY", "", CLR_SECTION_BG, 11, True),
        ("Total documents analysed", docs, None, 10, False),
        ("Total excerpts found", len(df), None, 10, False),
        ("Valid Evidence excerpts", valid_count, None, 10, False),
        ("Weak Evidence excerpts", weak_count, None, 10, False),
        ("Not Relevant excerpts", len(df) - valid_count - weak_count, None, 10, False),
        ("Skill themes detected", themes, None, 10, False),
        ("", "", None, 10, False),
        ("DOCUMENTS ANALYSED", "", CLR_SECTION_BG, 11, True),
    ]

    for label, value, bg, size, bold in cover_rows:
        row_idx = ws_cover.max_row + 1
        cell_a = ws_cover.cell(row=row_idx, column=1, value=label)
        cell_b = ws_cover.cell(row=row_idx, column=2, value=value if value != "" else None)
        cell_a.font = Font(name="Calibri", size=size, bold=bold,
                           color=CLR_HEADER_FG if bg == CLR_HEADER_BG else "000000")
        cell_b.font = body_font(size=size, bold=bold)
        if bg:
            cell_a.fill = fill(bg)
            cell_b.fill = fill(bg)
        cell_a.alignment = wrap_align()
        ws_cover.row_dimensions[row_idx].height = 22 if size >= 14 else 16

    for doc_name in sorted(df["Document"].unique()):
        doc_valid = len(df[(df["Document"] == doc_name) &
                           (df["Validation Category"] == "VALID EVIDENCE")])
        doc_weak  = len(df[(df["Document"] == doc_name) &
                           (df["Validation Category"] == "WEAK EVIDENCE")])
        row_idx = ws_cover.max_row + 1
        ws_cover.cell(row=row_idx, column=1, value=doc_name).font = body_font()
        ws_cover.cell(row=row_idx, column=2,
                      value=f"{doc_valid} valid, {doc_weak} weak").font = body_font()

    # ---- Theme pivot table on cover ----------------------------------------
    row_idx = ws_cover.max_row + 2
    ws_cover.cell(row=row_idx, column=1, value="THEME BREAKDOWN").font = \
        Font(name="Calibri", size=11, bold=True)
    ws_cover.cell(row=row_idx, column=1).fill = fill(CLR_SECTION_BG)
    row_idx += 1

    if not relevant.empty:
        pivot = (relevant.groupby(["Theme", "Validation Category"])
                 .size().unstack(fill_value=0).reset_index())
        headers = list(pivot.columns)
        for col_i, h in enumerate(headers, 1):
            c = ws_cover.cell(row=row_idx, column=col_i, value=str(h))
            c.font = header_font(size=10)
            c.fill = fill(CLR_HEADER_BG)
        row_idx += 1
        for _, prow in pivot.iterrows():
            for col_i, val in enumerate(prow, 1):
                ws_cover.cell(row=row_idx, column=col_i, value=val).font = body_font()
            row_idx += 1

    # ---- One sheet per document -------------------------------------------
    for doc_name in sorted(relevant["Document"].unique()):
        doc_df = relevant[relevant["Document"] == doc_name].sort_values(
            "Final Score", ascending=False
        )

        # Sheet name max 31 chars, no special characters
        safe_name = doc_name[:28].replace("/", "-").replace("\\", "-") \
                                 .replace("*", "").replace("?", "") \
                                 .replace("[", "").replace("]", "") \
                                 .replace(":", "-")
        ws = wb.create_sheet(title=safe_name)

        # Column widths
        ws.column_dimensions["A"].width = 22   # Theme
        ws.column_dimensions["B"].width = 7    # Page
        ws.column_dimensions["C"].width = 10   # Score
        ws.column_dimensions["D"].width = 16   # Category
        ws.column_dimensions["E"].width = 80   # Excerpt
        ws.column_dimensions["F"].width = 14   # Match type
        ws.column_dimensions["G"].width = 20   # Human label

        # Document title row
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value = f"Evidence Report — {doc_name}"
        title_cell.font = Font(name="Calibri", size=12, bold=True, color=CLR_HEADER_FG)
        title_cell.fill = fill(CLR_HEADER_BG)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24

        # Column headers
        headers = ["Theme", "Page", "Score", "Validation", "Excerpt", "Match Type", "Human Label"]
        for col_i, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=col_i, value=h)
            c.font = header_font(size=10)
            c.fill = fill(CLR_HEADER_BG)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin
        ws.row_dimensions[2].height = 18

        # Freeze panes below header
        ws.freeze_panes = "A3"

        # Data rows
        for row_i, (_, row) in enumerate(doc_df.iterrows(), start=3):
            is_valid = row["Validation Category"] == "VALID EVIDENCE"
            row_bg = CLR_VALID_BG if is_valid else CLR_WEAK_BG
            badge_color = CLR_VALID_BADGE if is_valid else CLR_WEAK_BADGE

            values = [
                row["Theme"],
                int(row["Page"]) if pd.notna(row["Page"]) else "",
                round(float(row["Final Score"]), 3),
                row["Validation Category"],
                row["Excerpt"],
                "keyword" if row["Keyword Hit"] else "semantic",
                row["Human Label"] if row["Human Label"] else "",
            ]
            for col_i, val in enumerate(values, 1):
                c = ws.cell(row=row_i, column=col_i, value=val)
                c.fill = fill(row_bg if row_i % 2 == 0 else CLR_ROW_ALT
                              if not is_valid else CLR_VALID_BG)
                c.border = thin
                if col_i == 5:  # Excerpt — wrap text
                    c.alignment = wrap_align()
                    c.font = body_font(size=9)
                elif col_i == 4:  # Category badge
                    c.font = body_font(size=9, bold=True, color=badge_color)
                    c.alignment = Alignment(horizontal="center", vertical="top")
                elif col_i == 3:  # Score
                    c.font = body_font(size=9, bold=True)
                    c.alignment = Alignment(horizontal="center", vertical="top")
                    c.number_format = "0.000"
                else:
                    c.font = body_font(size=9)
                    c.alignment = wrap_align()

            # Auto-height: ~15pt per 80 chars of excerpt
            excerpt_len = len(str(row["Excerpt"]))
            ws.row_dimensions[row_i].height = max(30, min(120, excerpt_len // 5))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# PDF evidence report
# ---------------------------------------------------------------------------

def to_pdf_report_bytes(results: List[SearchResult]) -> bytes:
    """
    Generates a formatted PDF evidence report using reportlab.
    Groups excerpts by document and theme.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable, PageBreak
        )
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
    except ImportError:
        raise ImportError(
            "reportlab is required for PDF export. "
            "Add 'reportlab' to requirements.txt."
        )

    df = results_to_dataframe(results)
    relevant = df[df["Validation Category"].isin(["VALID EVIDENCE", "WEAK EVIDENCE"])]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()

    # Custom styles
    style_title = ParagraphStyle(
        "ReportTitle", parent=styles["Title"],
        fontSize=18, textColor=colors.HexColor("#0A3D5C"),
        spaceAfter=6,
    )
    style_subtitle = ParagraphStyle(
        "Subtitle", parent=styles["Normal"],
        fontSize=10, textColor=colors.grey, spaceAfter=12,
    )
    style_doc_heading = ParagraphStyle(
        "DocHeading", parent=styles["Heading1"],
        fontSize=13, textColor=colors.HexColor("#0A3D5C"),
        spaceBefore=14, spaceAfter=4,
        borderPad=4,
    )
    style_theme_heading = ParagraphStyle(
        "ThemeHeading", parent=styles["Heading2"],
        fontSize=11, textColor=colors.HexColor("#6c5ce7"),
        spaceBefore=8, spaceAfter=3,
    )
    style_valid = ParagraphStyle(
        "ValidBadge", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#155724"),
        backColor=colors.HexColor("#D4EDDA"),
        borderPad=3, spaceAfter=2,
    )
    style_weak = ParagraphStyle(
        "WeakBadge", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#856404"),
        backColor=colors.HexColor("#FFF3CD"),
        borderPad=3, spaceAfter=2,
    )
    style_excerpt = ParagraphStyle(
        "Excerpt", parent=styles["Normal"],
        fontSize=9, leading=13,
        leftIndent=10, rightIndent=10,
        spaceAfter=6, textColor=colors.HexColor("#212529"),
    )
    style_meta = ParagraphStyle(
        "Meta", parent=styles["Normal"],
        fontSize=8, textColor=colors.grey,
        spaceAfter=8,
    )
    style_summary = ParagraphStyle(
        "Summary", parent=styles["Normal"],
        fontSize=10, leading=14, spaceAfter=4,
    )

    story = []
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    valid_count = len(df[df["Validation Category"] == "VALID EVIDENCE"])
    weak_count  = len(df[df["Validation Category"] == "WEAK EVIDENCE"])

    # Cover
    story.append(Paragraph("EO Skills Policy Evidence Report", style_title))
    story.append(Paragraph(f"Generated: {ts}", style_subtitle))
    story.append(HRFlowable(width="100%", thickness=1,
                            color=colors.HexColor("#0A3D5C"), spaceAfter=12))

    # Summary table
    summary_data = [
        ["Metric", "Count"],
        ["Documents analysed", str(df["Document"].nunique())],
        ["Total excerpts", str(len(df))],
        ["Valid Evidence", str(valid_count)],
        ["Weak Evidence", str(weak_count)],
        ["Not Relevant", str(len(df) - valid_count - weak_count)],
        ["Themes detected", str(df["Theme"].nunique())],
    ]
    summary_table = Table(summary_data, colWidths=[10*cm, 4*cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A3D5C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#F8F9FA"), colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.5*cm))

    if relevant.empty:
        story.append(Paragraph(
            "No valid or weak evidence found in this analysis run.",
            style_summary
        ))
        doc.build(story)
        buf.seek(0)
        return buf.read()

    story.append(PageBreak())

    # Evidence by document
    for doc_name in sorted(relevant["Document"].unique()):
        doc_df = relevant[relevant["Document"] == doc_name].sort_values(
            "Final Score", ascending=False
        )
        story.append(Paragraph(doc_name, style_doc_heading))
        story.append(HRFlowable(width="100%", thickness=0.5,
                                color=colors.HexColor("#CCCCCC"), spaceAfter=4))

        for theme in sorted(doc_df["Theme"].unique()):
            theme_df = doc_df[doc_df["Theme"] == theme]
            story.append(Paragraph(theme, style_theme_heading))

            for _, row in theme_df.iterrows():
                is_valid = row["Validation Category"] == "VALID EVIDENCE"
                badge_style = style_valid if is_valid else style_weak
                badge = "✓ VALID EVIDENCE" if is_valid else "⚠ WEAK EVIDENCE"

                story.append(Paragraph(badge, badge_style))
                story.append(Paragraph(
                    f"Page {int(row['Page'])}  |  "
                    f"Score: {row['Final Score']:.3f}  |  "
                    f"{'Keyword match' if row['Keyword Hit'] else 'Semantic match'}"
                    + (f"  |  Keyword: {row['Matched Keyword']}"
                       if row.get('Matched Keyword') else ""),
                    style_meta
                ))
                # Escape HTML special chars in excerpt
                excerpt_safe = (str(row["Excerpt"])
                                .replace("&", "&amp;")
                                .replace("<", "&lt;")
                                .replace(">", "&gt;"))
                story.append(Paragraph(excerpt_safe, style_excerpt))

                if row.get("Human Label"):
                    story.append(Paragraph(
                        f"Human review: {row['Human Label']}", style_meta
                    ))
                story.append(Spacer(1, 0.15*cm))

        story.append(PageBreak())

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# JSON export (D3-compatible)
# ---------------------------------------------------------------------------

def to_d3_json(results: List[SearchResult]) -> str:
    df = results_to_dataframe(results)
    relevant = df[df["Validation Category"].isin(["VALID EVIDENCE", "WEAK EVIDENCE"])]

    heatmap = []
    for _, row in relevant.iterrows():
        heatmap.append({
            "doc": row["Document"],
            "theme": row["Theme"],
            "score": round(float(row["Final Score"]), 3),
            "page": int(row["Page"]) if pd.notna(row["Page"]) else 0,
        })

    doc_nodes = [{"id": d, "group": "document"} for d in relevant["Document"].unique()]
    theme_nodes = [{"id": t, "group": "theme"} for t in relevant["Theme"].unique()]
    links = [
        {"source": row["Document"], "target": row["Theme"],
         "value": round(float(row["Final Score"]), 3)}
        for _, row in relevant.iterrows()
    ]

    output = {
        "generated_at": datetime.now().isoformat(),
        "heatmap_data": heatmap,
        "network_data": {"nodes": doc_nodes + theme_nodes, "links": links},
    }
    return json.dumps(output, indent=2)


# ---------------------------------------------------------------------------
# Annotated evidence report (Markdown)
# ---------------------------------------------------------------------------

def to_markdown_report(
    results: List[SearchResult],
    title: str = "EO Skills Policy Evidence Report",
) -> str:
    df = results_to_dataframe(results)
    lines: List[str] = []

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    lines.append(f"# {title}")
    lines.append(f"_Generated: {ts}_\n")
    lines.append(f"**Total evidence excerpts analysed:** {len(df)}")

    valid = df[df["Validation Category"] == "VALID EVIDENCE"]
    weak  = df[df["Validation Category"] == "WEAK EVIDENCE"]
    lines.append(f"- Valid Evidence: **{len(valid)}**")
    lines.append(f"- Weak Evidence: **{len(weak)}**")
    lines.append(f"- Not Relevant: **{len(df) - len(valid) - len(weak)}**\n")
    lines.append("---\n")

    for doc in df["Document"].unique():
        doc_df = df[df["Document"] == doc]
        lines.append(f"## {doc}")
        lines.append(
            f"_Pages: {int(doc_df['Page'].min())}–{int(doc_df['Page'].max())}_\n"
        )
        for theme in doc_df["Theme"].unique():
            theme_df = doc_df[doc_df["Theme"] == theme]
            rel = theme_df[theme_df["Validation Category"].isin(
                ["VALID EVIDENCE", "WEAK EVIDENCE"]
            )]
            if rel.empty:
                continue
            lines.append(f"### {theme}")
            for _, row in rel.sort_values("Final Score", ascending=False).iterrows():
                badge = "✅ VALID" if row["Validation Category"] == "VALID EVIDENCE" \
                    else "⚠️ WEAK"
                lines.append(
                    f"**[{badge}]** Score: `{row['Final Score']:.3f}` | "
                    f"Page {int(row['Page'])} | "
                    f"{'🔑 keyword' if row['Keyword Hit'] else '🔍 semantic'}"
                )
                lines.append(f"> {row['Excerpt']}\n")
                if row["Human Label"]:
                    lines.append(f"_Human review: {row['Human Label']}_\n")
        lines.append("---\n")

    return "\n".join(lines)
